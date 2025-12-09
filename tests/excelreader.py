# Excel reading and updating
from _datetime import datetime
import openpyxl

class ExcelDataProvider:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def get_test_data(self):
        test_data = []
        for row in range(2, self.sheet.max_row + 1):  # Skip header row
            test_id = self.sheet.cell(row=row, column=1).value
            username = self.sheet.cell(row=row, column=2).value
            password = self.sheet.cell(row=row, column=3).value
            test_data.append((test_id, username, password)) # [(test_id,username,password)]
        return test_data

    def write_result(self, test_id, result):
        date = datetime.today()
        time = datetime.now().strftime("%H:%M:%S")
        testername = "Nithya"
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == test_id:
                self.sheet.cell(row=row, column=4).value = result
                self.sheet.cell(row=row, column=5).value =date
                self.sheet.cell(row=row, column=6).value =time
                self.sheet.cell(row=row, column=7).value =testername
                break
        self.workbook.save(self.file_path)

    def close(self):
        self.workbook.close()
